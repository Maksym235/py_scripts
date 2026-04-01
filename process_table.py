from math import e

import openpyxl


def load_employees(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    assert ws is not None

    employees = []

    for row in ws.iter_rows(values_only=True):
        if row[0] == "Ім'я":
            continue

        employees.append(
            {"name": row[0], "department": row[1], "salary": row[2], "city": row[3]}
        )

    return employees


def filter_employees(employees, dept=None, city=None, min_salary=None):
    result = []

    for emp in employees:
        if dept and emp["department"] != dept:
            continue
        if city and emp["city"] != city:
            continue
        if min_salary and emp["salary"] < min_salary:
            continue

        result.append(emp)

    return result


def avg_salary(employees):
    if not employees:
        return 0

    total = sum(emp["salary"] for emp in employees)
    return total / len(employees)


def save_results(employees, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    ws.append(["Ім'я", "Відділ", "Зарплата", "Місто"])

    for emp in employees:
        ws.append([emp["name"], emp["department"], emp["salary"], emp["city"]])

    wb.save(filepath)
    print(f"✓ Збережено у {filepath}")


employees = load_employees("employees.xlsx")

print(f"Всього працівників: {len(employees)}")
print(f"Середня зарплата: {avg_salary(employees):.0f} грн\n")

# Фільтруємо — тільки IT відділ
it_team = filter_employees(employees, dept="IT")
print(f"IT відділ ({len(it_team)} чол.):")
for emp in it_team:
    print(f"  {emp['name']} — {emp['salary']} грн, {emp['city']}")

print(f"  Середня зарплата: {avg_salary(it_team):.0f} грн\n")

# Фільтруємо — Київ + зарплата від 80000
kyiv_senior = filter_employees(employees, city="Київ", min_salary=80000)
print(f"Київ, зарплата від 80000 ({len(kyiv_senior)} чол.):")
for emp in kyiv_senior:
    print(f"  {emp['name']} — {emp['salary']} грн, {emp['department']}")

# Зберігаємо результат
save_results(it_team, "it_team.xlsx")
